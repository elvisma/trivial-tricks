#!/usr/bin/env python
# coding: utf-8

# In[1]:


import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import random


# In[2]:


url1='https://www.douban.com/search?q='
head={}
head['User-Agent'] = 'Mozilla/5.0 (Linux; Android 4.1.1; Nexus 7 Build/JRO03D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166  Safari/535.19'


# In[3]:


movie_name=[]
global rows
rows=212
global nums
nums=1


# In[4]:


def fomt_str(str):
    name0=re.findall(r"^(.*?)】",str)
    name1=re.findall(r"\((.*?)$",str)
    name2= re.findall(r"^(\d\d月\d\d日 )", str)
    str0=str.strip()
    
    if name0:
        str0 = str0.replace(name0[0], ' ')
    if name1:
        str0=str0.replace(name1[0],' ')
    if name2:
        str0 = str0.replace(name2[0], ' ')
    
    str0=str0.strip().strip('(').strip(']').strip()
    return str0
     # str0=str.replace(name0[0],' ')
    


# In[7]:


def get_tags(name):
    
    tags=[] # 存放的标签
    aim_list=[]  # 获取的搜索地址
    resp=requests.get(url1+name)
    bs=BeautifulSoup(resp.text, 'lxml')
    a_list=bs.find_all('a')
    
    for a in a_list:
        if a.text != '' and 'https://www.douban.com/link2/?url=https%3A%2F%2Fmovie.douban.com%2Fsubject'in a['href']:
            aim_list.append([a.text.strip(),a['href']])

    #取豆瓣搜索出现的第一步电影或者电视剧为主
    if aim_list:
        resp2=requests.get(aim_list[0][1])
        bs_2=BeautifulSoup(resp2.text,"lxml")
        span_list=bs_2.find_all('span',attrs={'property':"v:genre"})
        for tag in span_list:
             tags.append(tag.text)
    else:
        tags.append(' ')
    return  tags


# In[8]:


def read_move_name(path):
    wb=load_workbook(path)
    sheet1=wb.get_sheet_by_name("Sheet1")
    for i in sheet1["D"]:
       # print(i.value,end=" ")
        if i.value != '影片名称':
            move_name.append(i.value)
    return wb


# In[9]:


def write_move_tag(wb,tag):
    global rows
    rows=rows+1
    #wd = load_workbook(path)
    cl=8
    sheet1=wb.get_sheet_by_name("Sheet1")
    for tags in tag:
        sheet1.cell(row=rows,column=cl).value=tags
        cl=cl+1
    wb.save("user_order_information_1.xlsx")


# In[10]:


if __name__ == '__main__':
    nu = 0
    wb=read_move_name("user_order_information.xlsx")
    for move in move_name:
        global nums
        nums = nums + 1
        if nums >= 212:
            if move != '节目名称':
                move=fomt_str(move)
                #time.sleep(random.random()*4)
                if nu == 100:
                    nu = 0
                    time.sleep(1000)
                tag=get_tags(move)
                write_move_tag(wb,tag)
                print(move,tag)
                print(nums,"/",len(move_name))


# In[ ]:




